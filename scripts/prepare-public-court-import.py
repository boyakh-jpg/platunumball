#!/usr/bin/env python3
"""Normalize the nationwide public basketball-court workbook for RankBall.

The script never writes to Supabase. Naver reverse geocoding is opt-in because it
can consume a paid quota. Output rows remain blocked until coordinates have been
reverse geocoded and generic names have been resolved.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


SCHEMA_VERSION = 1
NAVER_REVERSE_GEOCODE_URL = "https://maps.apigw.ntruss.com/map-reversegeocode/v2/gc"
REQUIRED_HEADERS = {
    "id",
    "구장명",
    "주소",
    "위도",
    "경도",
    "운영상태",
    "데이터출처",
    "출처ID",
    "출처라이선스",
    "이름보정대상",
}
ACCEPTED_NAME_STATUSES = {"확정", "완료", "승인", "approved", "verified", "complete", "completed"}
GENERIC_NAME_RE = re.compile(r"^(?:농구장|농구\s*코트|이름\s*없는\s*농구장)(?:\s*\([^)]*\))?$", re.IGNORECASE)
PHONE_RE = re.compile(r"^[0-9+()\-.,\s/]+$")

PROVIDER_MAP = {
    "OpenStreetMap": "openstreetmap",
    "전국체육시설 API": "sports_facility_api",
    "공공시설개방 표준데이터": "public_facility_open_data",
}
DATASET_MAP = {
    "openstreetmap": "openstreetmap",
    "sports_facility_api": "sports_facility_api",
    "public_facility_open_data": "public_facility_open_data",
}
LICENSE_MAP = {
    "openstreetmap": "OpenStreetMap ODbL 1.0",
    "sports_facility_api": "공공데이터포털 이용허락 제한 없음",
    "public_facility_open_data": "공공데이터포털 전국공공시설개방정보표준데이터",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="RankBall 공공 농구장 원장 정규화 및 드라이런 준비")
    parser.add_argument("--source", required=True, help="원본 .zip 또는 .xlsx")
    parser.add_argument("--output", required=True, help="정규화 JSON 출력")
    parser.add_argument("--report", required=True, help="요약 JSON 출력")
    parser.add_argument("--reverse-geocode", action="store_true", help="네이버 좌표 역지오코딩 실행")
    parser.add_argument("--env-file", help="네이버 키가 있는 dotenv 파일")
    parser.add_argument("--cache", help="역지오코딩 재개용 JSON 캐시")
    parser.add_argument("--geocode-limit", type=int, default=0, help="이번 실행의 최대 신규 API 호출 수, 0은 무제한")
    parser.add_argument("--delay-ms", type=int, default=100, help="네이버 요청 사이 대기 시간")
    return parser.parse_args()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def parse_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    normalized = clean_text(value).lower()
    if normalized in {"true", "t", "yes", "y", "1", "예", "있음", "유료"}:
        return True
    if normalized in {"false", "f", "no", "n", "0", "아니오", "없음", "무료"}:
        return False
    return None


def parse_float(value: Any) -> float | None:
    try:
        number = float(clean_text(value))
    except (TypeError, ValueError):
        return None
    return number if number == number and number not in (float("inf"), float("-inf")) else None


def parse_positive_int(value: Any) -> int | None:
    number = parse_float(value)
    if number is None or not number.is_integer() or number < 1 or number > 100:
        return None
    return int(number)


def parse_positive_float(value: Any) -> float | None:
    number = parse_float(value)
    return number if number is not None and number > 0 else None


def is_korea_coordinate(lat: float | None, lng: float | None) -> bool:
    return lat is not None and lng is not None and 33 <= lat <= 39.5 and 124 <= lng <= 132


def normalize_https_url(value: Any) -> str | None:
    raw = clean_text(value)
    if not raw:
        return None
    if re.match(r"^(?:www\.)?[a-z0-9.-]+\.[a-z]{2,}(?:/|$)", raw, re.IGNORECASE):
        raw = f"https://{raw}"
    try:
        parsed = urllib.parse.urlparse(raw)
    except ValueError:
        return None
    if parsed.scheme.lower() != "https" or not parsed.netloc:
        return None
    return urllib.parse.urlunparse(parsed)


def normalize_http_url(value: Any) -> str | None:
    raw = clean_text(value)
    if not raw:
        return None
    try:
        parsed = urllib.parse.urlparse(raw)
    except ValueError:
        return None
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return None
    return urllib.parse.urlunparse(parsed)


def load_dotenv(path: Path | None) -> None:
    if path is None:
        return
    for line in path.read_text(encoding="utf-8-sig").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def get_naver_credentials() -> tuple[str, str]:
    client_id = next(
        (
            clean_text(os.environ.get(key))
            for key in (
                "NAVER_MAP_CLIENT_ID",
                "NAVER_MAP_NCP_KEY_ID",
                "VITE_NAVER_MAP_CLIENT_ID",
                "VITE_NAVER_MAP_NCP_KEY_ID",
            )
            if clean_text(os.environ.get(key))
        ),
        "",
    )
    client_secret = next(
        (
            clean_text(os.environ.get(key))
            for key in ("NAVER_MAP_CLIENT_SECRET", "NAVER_MAP_NCP_KEY", "NAVER_MAP_NCP_CLIENT_SECRET")
            if clean_text(os.environ.get(key))
        ),
        "",
    )
    return client_id, client_secret


def open_source_workbook(source_path: Path) -> tuple[Any, str, str]:
    source_bytes = source_path.read_bytes()
    source_sha256 = hashlib.sha256(source_bytes).hexdigest()
    workbook_bytes = source_bytes
    workbook_name = source_path.name
    if source_path.suffix.lower() == ".zip":
        with zipfile.ZipFile(io.BytesIO(source_bytes)) as archive:
            workbook_entries = [name for name in archive.namelist() if name.lower().endswith(".xlsx") and not name.startswith("__MACOSX/")]
            if len(workbook_entries) != 1:
                raise ValueError(f"zip_xlsx_count_invalid:{len(workbook_entries)}")
            workbook_name = Path(workbook_entries[0]).name
            workbook_bytes = archive.read(workbook_entries[0])
    elif source_path.suffix.lower() != ".xlsx":
        raise ValueError("source_must_be_zip_or_xlsx")
    workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    return workbook, workbook_name, source_sha256


def sheet_rows(worksheet: Any) -> tuple[list[str], list[dict[str, Any]]]:
    iterator = worksheet.iter_rows(values_only=True)
    raw_headers = next(iterator, None)
    if raw_headers is None:
        return [], []
    headers = [clean_text(value) for value in raw_headers]
    rows: list[dict[str, Any]] = []
    for values in iterator:
        if not any(value is not None and clean_text(value) for value in values):
            continue
        rows.append({header: json_value(values[index] if index < len(values) else None) for index, header in enumerate(headers) if header})
    return headers, rows


def find_sheet(workbook: Any, name: str, fallback_index: int | None = None) -> Any | None:
    if name in workbook.sheetnames:
        return workbook[name]
    if fallback_index is not None and len(workbook.worksheets) > fallback_index:
        return workbook.worksheets[fallback_index]
    return None


def load_payload_map(workbook: Any) -> dict[str, dict[str, Any]]:
    worksheet = find_sheet(workbook, "RankBall_적재형식", 1)
    if worksheet is None:
        return {}
    _, rows = sheet_rows(worksheet)
    payloads: dict[str, dict[str, Any]] = {}
    for row in rows:
        court_id = clean_text(row.get("id"))
        raw_payload = row.get("payload")
        if not court_id or not isinstance(raw_payload, str):
            continue
        try:
            parsed = json.loads(raw_payload)
        except json.JSONDecodeError:
            continue
        if isinstance(parsed, dict):
            payloads[court_id] = parsed
    return payloads


def load_name_corrections(workbook: Any) -> dict[str, dict[str, str]]:
    worksheet = find_sheet(workbook, "네이버_이름대상", 2)
    if worksheet is None:
        return {}
    _, rows = sheet_rows(worksheet)
    corrections: dict[str, dict[str, str]] = {}
    for row in rows:
        court_id = clean_text(row.get("id"))
        corrected_name = clean_text(row.get("보정구장명"))
        if not court_id or not corrected_name:
            continue
        corrections[court_id] = {
            "name": corrected_name,
            "source": clean_text(row.get("이름출처")),
            "status": clean_text(row.get("이름검수상태")),
        }
    return corrections


def load_sfms_records(workbook: Any) -> dict[str, dict[str, Any]]:
    worksheet = find_sheet(workbook, "체육시설API_원본", 8)
    if worksheet is None:
        return {}
    _, rows = sheet_rows(worksheet)
    return {clean_text(row.get("faci_cd")).upper(): row for row in rows if clean_text(row.get("faci_cd"))}


def map_indoor_outdoor(value: Any) -> str:
    normalized = clean_text(value).lower()
    if normalized in {"야외", "실외", "outdoor"}:
        return "outdoor"
    if normalized in {"실내", "indoor"}:
        return "indoor"
    if normalized in {"실내외", "혼합", "mixed"}:
        return "mixed"
    return "unknown"


def map_venue_type(value: Any) -> str:
    normalized = clean_text(value).lower()
    if "공원" in normalized:
        return "park"
    if "학교" in normalized or "대학" in normalized:
        return "school"
    if "아파트" in normalized or "주거" in normalized:
        return "apartment"
    if "체육" in normalized or "민간" in normalized or "센터" in normalized:
        return "sports_facility"
    if "공공" in normalized:
        return "public_facility"
    return "unknown"


def map_surface(value: Any, indoor_outdoor: str) -> str:
    normalized = clean_text(value).lower()
    if not normalized or normalized in {"미확인", "확인 필요", "unknown"}:
        return "unknown"
    if "아스팔트" in normalized:
        return "asphalt"
    if any(token in normalized for token in ("우레탄", "탄성", "고무")):
        return "indoor_synthetic" if indoor_outdoor == "indoor" else "urethane"
    if any(token in normalized for token in ("흙", "마사", "모래", "클레이")):
        return "dirt"
    if any(token in normalized for token in ("마루", "목재", "원목")):
        return "indoor_wood"
    if indoor_outdoor == "indoor" and any(token in normalized for token in ("합성", "pvc", "synthetic")):
        return "indoor_synthetic"
    return "unknown"


def map_layout(value: Any) -> str:
    normalized = clean_text(value).lower()
    if not normalized or normalized in {"미확인", "확인 필요", "unknown"}:
        return "unknown"
    if "풀" in normalized or "full" in normalized:
        return "full"
    if "하프" in normalized or "반코트" in normalized or "half" in normalized:
        return "half"
    if re.search(r"(?:골대\s*1|1\s*개|단일)", normalized):
        return "single_hoop"
    return "unknown"


def map_access(value: Any) -> str:
    normalized = clean_text(value).lower()
    if any(token in normalized for token in ("공개", "자유", "walk_in", "open")):
        return "walk_in"
    if any(token in normalized for token in ("조건부", "예약", "reservation")):
        return "reservation"
    if any(token in normalized for token in ("제한", "비공개", "restricted")):
        return "restricted"
    return "unknown"


def map_operational_status(value: Any) -> str:
    normalized = clean_text(value).lower()
    if normalized in {"active", "정상", "정상운영", "운영"}:
        return "active"
    if normalized in {"pending", "대기", "검토"}:
        return "pending"
    if normalized in {"closed", "폐업", "폐쇄", "운영종료"}:
        return "closed"
    return "unknown"


def split_sfms_codes(value: Any, payload: dict[str, Any]) -> list[str]:
    candidates: list[Any] = []
    if isinstance(payload.get("sfmsFacilityCodes"), list):
        candidates.extend(payload["sfmsFacilityCodes"])
    raw = clean_text(value)
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                candidates.extend(parsed)
            else:
                candidates.append(parsed)
        except json.JSONDecodeError:
            candidates.extend(re.split(r"[,;|\s]+", raw))
    return sorted({clean_text(code).upper() for code in candidates if clean_text(code)})


def get_primary_provider(source_record_id: str) -> str:
    if source_record_id.startswith("osm:"):
        return "openstreetmap"
    if source_record_id.startswith("sfms:"):
        return "sports_facility_api"
    if source_record_id.startswith("public_data_portal:"):
        return "public_facility_open_data"
    return "unknown"


def provider_url(provider: str, urls: list[str], map_url: str) -> str | None:
    patterns = {
        "openstreetmap": "openstreetmap.org",
        "sports_facility_api": "/15096288/",
        "public_facility_open_data": "/15013117/",
    }
    pattern = patterns.get(provider, "")
    for url in urls:
        if pattern and pattern in url:
            return normalize_http_url(url)
    if provider == "openstreetmap":
        return normalize_http_url(map_url)
    return None


def source_record_id(provider: str, raw_record_id: str) -> str:
    return "court_source_" + hashlib.sha256(f"{provider}|{raw_record_id}".encode("utf-8")).hexdigest()[:24]


def normalize_source_records(
    row: dict[str, Any],
    payload: dict[str, Any],
    sfms_records: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    source_label = clean_text(row.get("데이터출처"))
    declared_providers = [PROVIDER_MAP.get(part.strip(), "unknown") for part in source_label.split("+") if part.strip()]
    primary_record_id = clean_text(row.get("출처ID"))
    primary_provider = get_primary_provider(primary_record_id)
    urls = [url for url in (normalize_http_url(item) for item in payload.get("sourceUrls", [])) if url]
    map_url = clean_text(row.get("지도링크"))
    sfms_codes = split_sfms_codes(row.get("SFMS시설코드"), payload)
    record_specs: list[tuple[str, str, str | None]] = []
    if primary_record_id:
        record_specs.append((primary_provider, primary_record_id, None))
    for code in sfms_codes:
        record_id = f"sfms:{code}"
        if not any(provider == "sports_facility_api" and existing_id.lower() == record_id.lower() for provider, existing_id, _ in record_specs):
            record_specs.append(("sports_facility_api", record_id, code))

    records: list[dict[str, Any]] = []
    for provider, record_id, facility_code in record_specs:
        sfms = sfms_records.get((facility_code or record_id.removeprefix("sfms:")).upper(), {}) if provider == "sports_facility_api" else {}
        external_code = facility_code or (record_id.removeprefix("sfms:") if provider == "sports_facility_api" else None)
        source_url = provider_url(provider, urls, map_url)
        source_metadata = {
            "declaredProviders": declared_providers,
            "sourceLabel": source_label,
            "reviewPriority": clean_text(row.get("검수우선순위")) or None,
            "dedupeGroupId": clean_text(row.get("중복그룹")) or None,
            "dedupeDistanceM": parse_float(row.get("병합거리(m)")),
            "mapUrl": normalize_http_url(map_url),
            "sfmsFacilityType": clean_text(row.get("SFMS시설유형")) or None,
            "sfmsBusinessType": clean_text(row.get("SFMS업종")) or None,
            "sfmsRegistrationStatus": clean_text(row.get("SFMS등록상태")) or None,
            "sourceHomepageRaw": clean_text(row.get("홈페이지")) or None,
        }
        records.append(
            {
                "id": source_record_id(provider, record_id),
                "provider": provider,
                "datasetId": DATASET_MAP.get(provider, provider),
                "sourceRecordId": record_id,
                "sourceUrl": source_url,
                "sourceLicense": LICENSE_MAP.get(provider, clean_text(row.get("출처라이선스")) or None),
                "sourceReferenceDate": clean_text(row.get("데이터기준일")) or None,
                "sourceRegisteredAt": clean_text(sfms.get("reg_dt")) or None,
                "sourceUpdatedAt": clean_text(sfms.get("updt_dt")) or None,
                "confidence": parse_float(row.get("신뢰도")),
                "externalFacilityCode": external_code,
                "sourceMetadata": source_metadata,
                "rawPayload": sfms if sfms else row,
            }
        )

    unresolved = sorted({provider for provider in declared_providers if provider != "unknown" and provider not in {record["provider"] for record in records}})
    return records, unresolved


def region_value(result: dict[str, Any] | None, area: str) -> str:
    if not result:
        return ""
    return clean_text(((result.get("region") or {}).get(area) or {}).get("name"))


def reverse_result(response: dict[str, Any], name: str) -> dict[str, Any] | None:
    results = ((response.get("v2") or {}).get("results") or response.get("results") or [])
    return next((result for result in results if result.get("name") == name), None)


def format_reverse_address(result: dict[str, Any] | None, include_building: bool = False) -> str:
    if not result:
        return ""
    regions = [region_value(result, key) for key in ("area1", "area2", "area3", "area4")]
    land = result.get("land") or {}
    number = "-".join(filter(None, (clean_text(land.get("number1")), clean_text(land.get("number2")))))
    mountain = "산" if result.get("name") == "addr" and str(land.get("type")) == "2" else ""
    road = clean_text(land.get("name")) if result.get("name") == "roadaddr" else ""
    addition0 = land.get("addition0") or {}
    building = clean_text(addition0.get("value")) if include_building and addition0.get("type") == "building" else ""
    lot_number = " ".join(filter(None, (mountain, number)))
    return clean_text(" ".join(filter(None, (*regions, road, lot_number, building))))


def normalize_reverse_response(response: dict[str, Any], lat: float, lng: float) -> dict[str, Any]:
    status = (response.get("status") or (response.get("v2") or {}).get("status") or {})
    if str(status.get("code", "0")) != "0":
        raise ValueError(f"naver_reverse_status:{status.get('code')}")
    road_result = reverse_result(response, "roadaddr")
    jibun_result = reverse_result(response, "addr")
    legal_result = reverse_result(response, "legalcode")
    admin_result = reverse_result(response, "admcode")
    region_result = road_result or jibun_result or legal_result or admin_result
    address = (response.get("v2") or {}).get("address") or response.get("address") or {}
    road_address = clean_text(address.get("roadAddress") or format_reverse_address(road_result, True))
    jibun_address = clean_text(address.get("jibunAddress") or format_reverse_address(jibun_result))
    address_text = road_address or jibun_address
    if not address_text or not region_result:
        raise ValueError("naver_reverse_address_missing")
    land = (road_result or {}).get("land") or {}
    addition0 = land.get("addition0") or {}
    addition1 = land.get("addition1") or {}
    legal_or_region = jibun_result or legal_result or region_result
    admin_or_region = admin_result or region_result
    return {
        "addressText": address_text,
        "roadAddress": road_address or None,
        "jibunAddress": jibun_address or None,
        "buildingName": clean_text(addition0.get("value")) if addition0.get("type") == "building" else None,
        "sido": region_value(region_result, "area1"),
        "sigungu": region_value(region_result, "area2"),
        "emd": region_value(legal_or_region, "area4") or region_value(legal_or_region, "area3") or region_value(admin_or_region, "area3"),
        "zonecode": clean_text(addition1.get("value")) if addition1.get("type") == "zipcode" else None,
        "lat": lat,
        "lng": lng,
    }


def fetch_naver_reverse(lat: float, lng: float, client_id: str, client_secret: str) -> dict[str, Any]:
    query = urllib.parse.urlencode(
        {
            "coords": f"{lng:.10f},{lat:.10f}",
            "output": "json",
            "orders": "legalcode,admcode,addr,roadaddr",
        }
    )
    request = urllib.request.Request(
        f"{NAVER_REVERSE_GEOCODE_URL}?{query}",
        headers={
            "x-ncp-apigw-api-key-id": client_id,
            "x-ncp-apigw-api-key": client_secret,
            "Accept": "application/json",
            "User-Agent": "RankBallPublicCourtImporter/1.0",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        raise RuntimeError(f"naver_reverse_http_{error.code}") from error
    except urllib.error.URLError as error:
        raise RuntimeError(f"naver_reverse_network:{error.reason}") from error
    return normalize_reverse_response(payload, lat, lng)


def load_cache(path: Path | None) -> dict[str, Any]:
    if path is None or not path.exists():
        return {"schemaVersion": 1, "entries": {}}
    parsed = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(parsed, dict) or not isinstance(parsed.get("entries"), dict):
        raise ValueError("invalid_geocode_cache")
    return parsed


def save_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(path.suffix + ".tmp")
    temporary_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary_path.replace(path)


def make_hashtags(rows: list[dict[str, Any]]) -> dict[str, str]:
    used: set[str] = set()
    hashtags: dict[str, str] = {}
    for court_id in sorted(clean_text(row.get("id")) for row in rows if clean_text(row.get("id"))):
        number = 10000 + (int(hashlib.sha256(court_id.encode("utf-8")).hexdigest()[:12], 16) % 90000)
        while f"#{number:05d}" in used:
            number = 10000 + ((number - 9999) % 90000)
        hashtag = f"#{number:05d}"
        used.add(hashtag)
        hashtags[court_id] = hashtag
    return hashtags


def import_key_for(row: dict[str, Any]) -> str:
    material = {
        "court": row["court"],
        "facilityInfo": row["facilityInfo"],
        "sources": row["sources"],
    }
    encoded = json.dumps(material, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def issue_disposition(issues: list[str]) -> str:
    if "source_not_active" in issues:
        return "excluded"
    if "invalid_korea_coordinate" in issues or "source_record_required" in issues:
        return "invalid"
    if any(issue in issues for issue in ("generic_or_unresolved_name", "name_correction_unverified", "source_coordinate_collision")):
        return "review_required"
    if "reverse_geocode_failed" in issues:
        return "geocode_failed"
    if "reverse_geocode_required" in issues or "address_missing" in issues:
        return "needs_geocode"
    return "ready"


def main() -> int:
    args = parse_args()
    source_path = Path(args.source).resolve()
    output_path = Path(args.output).resolve()
    report_path = Path(args.report).resolve()
    cache_path = Path(args.cache).resolve() if args.cache else output_path.with_name("naver-reverse-geocode-cache.json")
    if not source_path.exists():
        raise FileNotFoundError(source_path)

    load_dotenv(Path(args.env_file).resolve() if args.env_file else None)
    workbook, workbook_name, source_sha256 = open_source_workbook(source_path)
    primary_sheet = find_sheet(workbook, "최종_농구장", 0)
    if primary_sheet is None:
        raise ValueError("primary_sheet_missing")
    headers, source_rows = sheet_rows(primary_sheet)
    missing_headers = sorted(REQUIRED_HEADERS - set(headers))
    if missing_headers:
        raise ValueError(f"required_headers_missing:{','.join(missing_headers)}")
    if not source_rows:
        raise ValueError("primary_sheet_has_no_data_rows")

    payload_map = load_payload_map(workbook)
    corrections = load_name_corrections(workbook)
    sfms_records = load_sfms_records(workbook)
    hashtags = make_hashtags(source_rows)

    coordinate_groups: dict[tuple[float, float], list[str]] = defaultdict(list)
    for row in source_rows:
        lat = parse_float(row.get("위도"))
        lng = parse_float(row.get("경도"))
        if is_korea_coordinate(lat, lng):
            coordinate_groups[(round(lat, 7), round(lng, 7))].append(clean_text(row.get("id")))
    collision_ids = {
        court_id
        for court_ids in coordinate_groups.values()
        if len(court_ids) > 1
        for court_id in court_ids
    }

    cache = load_cache(cache_path) if args.reverse_geocode else {"schemaVersion": 1, "entries": {}}
    cache_entries = cache["entries"]
    client_id = client_secret = ""
    if args.reverse_geocode:
        client_id, client_secret = get_naver_credentials()
        if not client_id or not client_secret:
            raise ValueError("naver_reverse_geocode_credentials_missing")

    geocode_stats = Counter()
    normalized_rows: list[dict[str, Any]] = []
    new_geocode_calls = 0
    generated_at = datetime.now(timezone.utc).isoformat()

    for worksheet_row_number, source_row in enumerate(source_rows, start=2):
        court_id = clean_text(source_row.get("id"))
        original_name = clean_text(source_row.get("구장명"))
        name = original_name
        name_source = "source"
        correction = corrections.get(court_id)
        correction_status = clean_text((correction or {}).get("status")).lower()
        correction_is_accepted = bool(correction and correction_status in ACCEPTED_NAME_STATUSES)
        if correction_is_accepted:
            name = clean_text(correction["name"])
            name_source = "naver_place" if "네이버" in clean_text(correction.get("source")) else "manual"

        lat = parse_float(source_row.get("위도"))
        lng = parse_float(source_row.get("경도"))
        valid_coordinate = is_korea_coordinate(lat, lng)
        operational_status = map_operational_status(source_row.get("운영상태"))
        rankball_payload = payload_map.get(court_id, {})
        sources, unresolved_providers = normalize_source_records(source_row, rankball_payload, sfms_records)
        sfms_codes = split_sfms_codes(source_row.get("SFMS시설코드"), rankball_payload)
        sfms_rows = [sfms_records[code] for code in sfms_codes if code in sfms_records]
        zonecode = next((clean_text(row.get("faci_road_zip") or row.get("faci_zip")) for row in sfms_rows if clean_text(row.get("faci_road_zip") or row.get("faci_zip"))), "")
        facility_area = next((parse_positive_float(row.get("faci_gfa")) for row in sfms_rows if parse_positive_float(row.get("faci_gfa")) is not None), None)
        indoor_outdoor = map_indoor_outdoor(source_row.get("실내외"))
        venue_type = map_venue_type(source_row.get("시설유형"))
        surface_raw = clean_text(source_row.get("바닥"))
        layout_raw = clean_text(source_row.get("코트규격"))
        access_type = map_access(source_row.get("접근구분"))
        homepage_raw = clean_text(source_row.get("홈페이지"))
        official_url = normalize_https_url(homepage_raw)
        contact_phone = clean_text(source_row.get("전화")) or None

        reverse_address: dict[str, Any] | None = None
        reverse_error: str | None = None
        reverse_verified_at: str | None = None
        if args.reverse_geocode and valid_coordinate and operational_status == "active":
            cache_key = f"{lng:.7f},{lat:.7f}"
            cached = cache_entries.get(cache_key)
            if isinstance(cached, dict) and isinstance(cached.get("address"), dict):
                reverse_address = cached["address"]
                reverse_verified_at = clean_text(cached.get("fetchedAt")) or None
                geocode_stats["cached"] += 1
            elif args.geocode_limit and new_geocode_calls >= args.geocode_limit:
                geocode_stats["deferred_by_limit"] += 1
            else:
                geocode_stats["attempted"] += 1
                new_geocode_calls += 1
                try:
                    reverse_address = fetch_naver_reverse(lat, lng, client_id, client_secret)
                    reverse_verified_at = datetime.now(timezone.utc).isoformat()
                    cache_entries[cache_key] = {"address": reverse_address, "fetchedAt": reverse_verified_at}
                    geocode_stats["succeeded"] += 1
                    save_json(cache_path, cache)
                except (RuntimeError, ValueError, json.JSONDecodeError) as error:
                    reverse_error = str(error)[:200]
                    geocode_stats["failed"] += 1
                if args.delay_ms > 0:
                    time.sleep(args.delay_ms / 1000)

        address_text = clean_text(source_row.get("주소"))
        road_address = clean_text(source_row.get("도로명주소")) or None
        jibun_address = clean_text(source_row.get("지번주소")) or None
        sido = clean_text(source_row.get("시도"))
        sigungu = clean_text(source_row.get("시군구"))
        emd = clean_text(source_row.get("읍면동"))
        address_source = "source"
        geocode_verified = False
        verified_at = None
        if reverse_address:
            address_text = clean_text(reverse_address.get("addressText"))
            road_address = clean_text(reverse_address.get("roadAddress")) or None
            jibun_address = clean_text(reverse_address.get("jibunAddress")) or None
            zonecode = clean_text(reverse_address.get("zonecode")) or zonecode
            sido = clean_text(reverse_address.get("sido"))
            sigungu = clean_text(reverse_address.get("sigungu"))
            emd = clean_text(reverse_address.get("emd"))
            address_source = "naver_reverse_geocode"
            geocode_verified = True
            verified_at = reverse_verified_at or generated_at

        issues: list[str] = []
        if operational_status != "active":
            issues.append("source_not_active")
        if not court_id:
            issues.append("court_id_missing")
        if not valid_coordinate:
            issues.append("invalid_korea_coordinate")
        name_requires_correction = parse_bool(source_row.get("이름보정대상")) is True or GENERIC_NAME_RE.match(name) is not None
        if name_requires_correction and not correction_is_accepted:
            issues.append("generic_or_unresolved_name")
        if correction and not correction_is_accepted:
            issues.append("name_correction_unverified")
        if court_id in collision_ids:
            issues.append("source_coordinate_collision")
        if not sources:
            issues.append("source_record_required")
        if unresolved_providers:
            issues.append("source_provider_without_record_id")
        if not address_text:
            issues.append("address_missing")
        if operational_status == "active" and valid_coordinate and not geocode_verified:
            issues.append("reverse_geocode_failed" if reverse_error else "reverse_geocode_required")
        if homepage_raw and official_url is None:
            issues.append("official_url_not_https")
        if contact_phone and not PHONE_RE.fullmatch(contact_phone):
            issues.append("contact_phone_format_review")

        reservation_required = True if access_type == "reservation" else False if access_type == "walk_in" else None
        court = {
            "id": court_id,
            "name": name,
            "hashtag": hashtags.get(court_id),
            "addressText": address_text or None,
            "roadAddress": road_address,
            "jibunAddress": jibun_address,
            "zonecode": zonecode or None,
            "lat": lat,
            "lng": lng,
            "regionKey": None,
            "registrationOrigin": "public_import",
            "facilityName": name,
            "courtUnit": None,
            "sido": sido or None,
            "sigungu": sigungu or None,
            "emd": emd or None,
            "indoorOutdoor": indoor_outdoor,
            "venueType": venue_type,
            "courtKind": "unknown",
            "surfaceType": map_surface(surface_raw, indoor_outdoor),
            "surfaceTypeRaw": surface_raw or None,
            "courtLayout": map_layout(layout_raw),
            "courtLayoutRaw": layout_raw or None,
            "hoopCount": parse_positive_int(source_row.get("골대수")),
            "accessType": access_type,
            "reservationRequired": reservation_required,
            "paid": parse_bool(source_row.get("유료")),
            "lighting": parse_bool(source_row.get("조명")),
            "operationalStatus": operational_status,
            "verificationStatus": "source_verified" if geocode_verified and not name_requires_correction else "pending",
            "nameSource": name_source,
            "addressSource": address_source,
            "sourceConfidence": parse_float(source_row.get("신뢰도")),
            "verifiedAt": verified_at,
            "geocodeVerified": geocode_verified,
            "multipleCourtsVerified": False,
            "payload": {
                "baseName": name,
                "facilityName": name,
                "courtUnit": None,
                "canonicalBaseName": name,
                "region": sigungu or sido or None,
                "addressDong": emd or None,
                "type": "야외" if indoor_outdoor == "outdoor" else "실내" if indoor_outdoor == "indoor" else "확인 필요",
                "courtKind": "unknown",
                "surfaceType": map_surface(surface_raw, indoor_outdoor),
                "courtLayout": map_layout(layout_raw),
                "accessType": access_type,
                "reservation": reservation_required,
                "paid": parse_bool(source_row.get("유료")),
                "lighting": parse_bool(source_row.get("조명")),
                "sourceUrl": official_url,
                "registrationOrigin": "public_import",
            },
        }
        facility_info = {
            "operatorName": clean_text(source_row.get("운영기관")) or None,
            "contactPhone": contact_phone,
            "officialUrl": official_url,
            "reservationUrl": None,
            "openingHoursText": clean_text(source_row.get("운영시간")) or None,
            "applicationMethod": clean_text(source_row.get("신청방법")) or None,
            "accessNote": clean_text(source_row.get("접근근거")) or None,
            "detailAddress": None,
            "locationNote": None,
            "facilityAreaSqm": facility_area,
            "facilityAreaScope": "facility" if facility_area is not None else None,
        }
        normalized_row = {
            "rowNumber": worksheet_row_number,
            "disposition": "",
            "issues": sorted(set(issues)),
            "court": court,
            "facilityInfo": facility_info,
            "sources": sources,
            "rawPayload": {
                "sheet": primary_sheet.title,
                "rowNumber": worksheet_row_number,
                "row": source_row,
                "rankBallLoadPayload": rankball_payload,
                "reverseGeocodeError": reverse_error,
                "unresolvedProviders": unresolved_providers,
            },
        }
        normalized_row["disposition"] = issue_disposition(normalized_row["issues"])
        if normalized_row["disposition"] == "ready":
            court["verificationStatus"] = "source_verified"
        normalized_row["importKey"] = import_key_for(normalized_row)
        court["payload"]["publicImportKey"] = normalized_row["importKey"]
        normalized_rows.append(normalized_row)

    disposition_counts = Counter(row["disposition"] for row in normalized_rows)
    issue_counts = Counter(issue for row in normalized_rows for issue in row["issues"])
    source_counts = Counter(source["provider"] for row in normalized_rows for source in row["sources"])
    field_coverage = {
        "operatorName": sum(bool(row["facilityInfo"]["operatorName"]) for row in normalized_rows),
        "contactPhone": sum(bool(row["facilityInfo"]["contactPhone"]) for row in normalized_rows),
        "officialUrlHttps": sum(bool(row["facilityInfo"]["officialUrl"]) for row in normalized_rows),
        "openingHours": sum(bool(row["facilityInfo"]["openingHoursText"]) for row in normalized_rows),
        "applicationMethod": sum(bool(row["facilityInfo"]["applicationMethod"]) for row in normalized_rows),
        "zonecode": sum(bool(row["court"]["zonecode"]) for row in normalized_rows),
        "facilityAreaSqm": sum(row["facilityInfo"]["facilityAreaSqm"] is not None for row in normalized_rows),
    }
    summary = {
        "totalRows": len(normalized_rows),
        "readyRows": disposition_counts.get("ready", 0),
        "blockedRows": len(normalized_rows) - disposition_counts.get("ready", 0),
        "dispositionCounts": dict(sorted(disposition_counts.items())),
        "issueCounts": dict(sorted(issue_counts.items())),
        "sourceRecordCounts": dict(sorted(source_counts.items())),
        "fieldCoverage": field_coverage,
        "exactCoordinateCollisionGroups": sum(len(ids) > 1 for ids in coordinate_groups.values()),
        "exactCoordinateCollisionRows": len(collision_ids),
        "geocode": {
            "requested": args.reverse_geocode,
            "newCallLimit": args.geocode_limit,
            **dict(sorted(geocode_stats.items())),
        },
    }
    batch_id = f"public-courts-{source_sha256[:16]}"
    output = {
        "schemaVersion": SCHEMA_VERSION,
        "batchId": batch_id,
        "source": {
            "archiveFileName": source_path.name,
            "workbookFileName": workbook_name,
            "sheetName": primary_sheet.title,
            "sha256": source_sha256,
            "generatedAt": generated_at,
        },
        "summary": summary,
        "rows": normalized_rows,
    }
    report = {
        "schemaVersion": SCHEMA_VERSION,
        "batchId": batch_id,
        "source": output["source"],
        "summary": summary,
        "samples": {
            disposition: [
                {
                    "rowNumber": row["rowNumber"],
                    "courtId": row["court"]["id"],
                    "name": row["court"]["name"],
                    "issues": row["issues"],
                }
                for row in normalized_rows
                if row["disposition"] == disposition
            ][:10]
            for disposition in sorted(disposition_counts)
        },
    }
    save_json(output_path, output)
    save_json(report_path, report)
    if args.reverse_geocode:
        save_json(cache_path, cache)
    print(json.dumps({"output": str(output_path), "report": str(report_path), "batchId": batch_id, "summary": summary}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        print(f"public_court_prepare_failed:{error}", file=sys.stderr)
        raise
